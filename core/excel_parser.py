# core/excel_parser.py  (pure openpyxl, two-block parser)
from __future__ import annotations
from io import BytesIO
import re
from typing import Dict, Any
from openpyxl import load_workbook

def sanitize_key(s: str) -> str:
    """Turn Excel text into a safe placeholder key."""
    return (
        str(s).strip()
        .replace(" ", "_").replace("/", "_").replace("\\", "_")
        .replace("-", "_").replace("(", "").replace(")", "")
        .replace(".", "_").replace(":", "_")
    )

def _fmt3(v):
    """Return a string with 3 decimals for numbers, or original if not numeric."""
    if v is None:
        return None
    try:
        return f"{float(str(v).strip()):.3f}"
    except Exception:
        return v

def list_sheet_names_bytes(file_bytes: bytes) -> list[str]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()

# Helpers for column letters
_LETTERS_RE = re.compile(r"^[A-Za-z]+$")

def _col_letter_to_index(letter: str) -> int:
    """A→0, B→1, …, Z→25, AA→26, etc."""
    s = str(letter).strip().upper()
    idx = 0
    for ch in s:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1

def parse_excel_two_blocks_bytes(
    file_bytes: bytes,
    sheet_name: str,
    # LEFT block (XYZ)
    left_node_col: str = "A",
    left_x_col: str = "B",
    left_y_col: str = "C",
    left_z_col: str = "D",
    left_has_header: bool = True,
    left_start_row: int | None = None,   # if None, computed from header flag
    # RIGHT block (key/value)
    right_key_col: str = "I",
    right_val_col: str = "J",
    right_has_header: bool = True,
    right_start_row: int | None = None,  # if None, computed from header flag
    # optional subset of node names (match on original A values)
    only_nodes: list[str] | None = None,
) -> Dict[str, Any]:
    """
    Parse a single sheet that has two areas:

    - LEFT block: A=Placeholder (node name), B/C/D= X/Y/Z
      -> produces keys like NodeName_X / _Y / _Z

    - RIGHT block: I=Placeholder (key), J=Value
      -> produces keys like IP_1, SN_SBG1, etc.

    All numeric values are formatted to 3 decimals (strings).
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
        ws = wb[sheet_name]

        out: Dict[str, Any] = {}
        seen_nodes: set[str] = set()
        only_set = set(str(n).strip() for n in only_nodes) if only_nodes else None

        # -------- LEFT block (XYZ) --------
        l_start = (2 if left_has_header else 1) if left_start_row is None else left_start_row

        ln = _col_letter_to_index(left_node_col)
        lx = _col_letter_to_index(left_x_col)
        ly = _col_letter_to_index(left_y_col)
        lz = _col_letter_to_index(left_z_col)

        for row in ws.iter_rows(min_row=l_start, values_only=True):
            if row is None or all(v is None for v in row):
                continue
            # guard missing columns
            if len(row) <= max(ln, lx, ly, lz):
                continue

            raw_node = row[ln]
            if raw_node is None:
                continue
            node_name = str(raw_node).strip()
            if not node_name:
                continue
            if only_set and node_name not in only_set:
                continue

            base = sanitize_key(node_name)
            key = base
            k = 2
            while key in seen_nodes:  # avoid collision on duplicate node names
                key = f"{base}_{k}"
                k += 1
            seen_nodes.add(key)

            out[f"{key}_X"] = _fmt3(row[lx])
            out[f"{key}_Y"] = _fmt3(row[ly])
            out[f"{key}_Z"] = _fmt3(row[lz])

        # -------- RIGHT block (key/value) --------
        r_start = (2 if right_has_header else 1) if right_start_row is None else right_start_row

        rk = _col_letter_to_index(right_key_col)
        rv = _col_letter_to_index(right_val_col)

        for row in ws.iter_rows(min_row=r_start, values_only=True):
            if row is None:
                continue
            if len(row) <= max(rk, rv):
                continue
            ph, val = row[rk], row[rv]
            if ph is None:
                continue
            key = sanitize_key(ph)
            out[key] = _fmt3(val)

        return out
    finally:
        wb.close()
